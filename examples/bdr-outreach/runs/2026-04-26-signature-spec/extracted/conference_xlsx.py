"""XLSX output builder for the conference enrichment path.

All formatting constants (colors, font sizes, column widths) are lifted
verbatim from references/conference-enrichment.md lines 151–174 (openpyxl
code block). The extract_linkedin helper is lifted from lines 107–113.

Source: references/conference-enrichment.md
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

# ── Formatting constants (from references/conference-enrichment.md) ──

HEADER_FILL_COLOR = "1F4E79"   # dark navy
ALT_ROW_FILL_COLOR = "EBF3FB"  # light blue
HEADER_FONT_SIZE = 12
DATA_FONT_SIZE = 11

COLUMNS = ["Company", "First Name", "Last Name", "Job Title", "LinkedIn", "Email"]
COLUMN_WIDTHS = [30, 18, 18, 35, 45, 35]


def extract_linkedin(external_urls: list[dict[str, Any]] | None) -> str:
    """Extract LinkedIn URL from ZoomInfo externalUrls field.

    Lifted verbatim from references/conference-enrichment.md lines 107–113.
    """
    if not external_urls:
        return ""
    for url_obj in external_urls:
        url = url_obj.get("url", "")
        if "linkedin.com" in url.lower():
            return url
    return ""


def build_conference_xlsx(
    enriched_contacts: list[dict[str, Any]],
    conference_name: str,
) -> dict[str, Any]:
    """Build a formatted XLSX from enriched conference contacts.

    Formatting from references/conference-enrichment.md:
    - Font: Arial 11 data / 12 bold white headers
    - Header fill: #1F4E79 (dark navy)
    - Alternating row fill: #EBF3FB (light blue) / white
    - Frozen pane at A2, auto-filter on header row
    - Column widths: Company=30, First/Last=18, Title=35, LinkedIn=45, Email=35
    - Output filename: YYYY_ConferenceName_Pharma_Contacts.xlsx

    Returns a dict with xlsx_path and coverage_summary.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl required. Run: pip install openpyxl")

    from datetime import datetime
    year = datetime.now().strftime("%Y")
    safe_name = conference_name.replace(" ", "_").replace("/", "-")
    output_path = Path(f"{year}_{safe_name}_Pharma_Contacts.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Pharma Contacts"

    header_fill = PatternFill(
        start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid"
    )
    alt_fill = PatternFill(
        start_color=ALT_ROW_FILL_COLOR, end_color=ALT_ROW_FILL_COLOR, fill_type="solid"
    )
    header_font = Font(name="Arial", size=HEADER_FONT_SIZE, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=DATA_FONT_SIZE)

    for col_idx, (header, width) in enumerate(zip(COLUMNS, COLUMN_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    email_count = 0
    linkedin_count = 0

    for row_idx, contact in enumerate(enriched_contacts, start=2):
        is_even = (row_idx % 2) == 0
        linkedin = extract_linkedin(contact.get("externalUrls"))
        email = contact.get("email", "")

        if email:
            email_count += 1
        if linkedin:
            linkedin_count += 1

        row_data = [
            contact.get("companyName", ""),
            contact.get("firstName", ""),
            contact.get("lastName", ""),
            contact.get("jobTitle", ""),
            linkedin,
            email,
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            if is_even:
                cell.fill = alt_fill

    wb.save(output_path)
    total = len(enriched_contacts)
    return {
        "xlsx_path": str(output_path),
        "coverage_summary": {
            "total": total,
            "email_count": email_count,
            "email_pct": round(email_count / total * 100, 1) if total else 0,
            "linkedin_count": linkedin_count,
            "linkedin_pct": round(linkedin_count / total * 100, 1) if total else 0,
        },
    }
