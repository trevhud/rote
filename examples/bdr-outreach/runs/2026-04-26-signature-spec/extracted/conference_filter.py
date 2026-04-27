"""Conference attendee list filtering functions.

The is_pharma() function and keyword lists are lifted verbatim from
references/conference-enrichment.md lines 52–85 (Python code block).
The load_attendee_file and partition_contacts functions codify the
fixed Steps 1–2 of the conference enrichment workflow.

Source: references/conference-enrichment.md
"""

from __future__ import annotations

from pathlib import Path
from typing import Any


# ── Keyword lists (lifted verbatim from references/conference-enrichment.md) ──

EXCLUDE_KEYWORDS = [
    "consulting",
    "consultancy",
    "advisory",
    "advisors",
    "partners llp",
    "university",
    "college",
    "institute",
    "hospital",
    "health system",
    "insurance",
    "iqvia",
    "parexel",
    "covance",
    "icon plc",
    "medidata",
    "veeva",
    "oracle",
    "microsoft",
    "google",
    "salesforce",
    "sas institute",
    "mckinsey",
    "deloitte",
    "accenture",
    "pwc",
    "bain & company",
    "bcg",
    "ernst & young",
    "kpmg",
    "government",
    "agency",
    "foundation",
]

INCLUDE_KEYWORDS = [
    "pharma",
    "biopharm",
    "biotech",
    "therapeutics",
    "biopharma",
    "biologics",
    "oncology",
    "genomics",
    "sciences",
    "medicines",
    "abbvie",
    "pfizer",
    "roche",
    "novartis",
    "merck",
    "lilly",
    "astrazeneca",
    "genentech",
    "amgen",
    "regeneron",
    "biogen",
    "gilead",
    "bristol",
    "sanofi",
    "gsk",
    "glaxo",
    "boehringer",
    "novo nordisk",
    "takeda",
    "bayer",
    "janssen",
    "astellas",
    "eisai",
    "daiichi",
    "otsuka",
    "lundbeck",
    "ucb",
    "vertex",
    "moderna",
    "alexion",
    "shire",
    "allergan",
    "ipsen",
]


def is_pharma(company_name: str) -> bool:
    """Classify a company as pharma/biotech (True) or not (False).

    Lifted verbatim from references/conference-enrichment.md lines 52–85.
    Exclude keywords take precedence over include keywords.
    Returns False (exclude) by default for ambiguous names.
    """
    if not company_name:
        return False
    name = company_name.lower()
    for kw in EXCLUDE_KEYWORDS:
        if kw in name:
            return False
    for kw in INCLUDE_KEYWORDS:
        if kw in name:
            return True
    return False  # default: exclude if unclear


def load_attendee_file(file_path: str) -> list[dict[str, Any]]:
    """Load a conference attendee CSV or XLSX into a normalized list of contact dicts.

    Minimum required columns: Company (or Organization), First Name, Last Name.
    Job Title is optional but used by downstream vetting.

    Returns a list of raw contact dicts with normalized keys.
    """
    path = Path(file_path)
    if path.suffix.lower() in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl required for XLSX input. Run: pip install openpyxl")
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h else "" for h in rows[0]]
        return [
            {headers[i]: (row[i] or "") for i in range(len(headers))}
            for row in rows[1:]
        ]
    elif path.suffix.lower() == ".csv":
        import csv
        with path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            return [dict(row) for row in reader]
    else:
        raise ValueError(f"Unsupported file format: {path.suffix}. Expected .csv or .xlsx")


def partition_contacts(
    raw_contacts: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Partition attendee list into pharma_contacts and excluded_contacts.

    The excluded list must be presented to the user via conf_exclusion_review_gate
    before enrichment begins — it is faster to add back a misclassified company
    now than to re-run ZoomInfo enrichment after the fact.

    Returns (pharma_contacts, excluded_contacts).
    """
    pharma: list[dict[str, Any]] = []
    excluded: list[dict[str, Any]] = []
    for c in raw_contacts:
        company = (
            c.get("Company")
            or c.get("company")
            or c.get("Organization")
            or c.get("organization", "")
        )
        if is_pharma(str(company)):
            pharma.append(c)
        else:
            excluded.append(c)
    return pharma, excluded
