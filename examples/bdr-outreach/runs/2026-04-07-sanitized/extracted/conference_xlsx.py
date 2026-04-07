"""
Extracted from: examples/bdr-outreach/skill/references/conference-enrichment.md
Lines 143–174 — the openpyxl XLSX builder.

The LLM was generating this formatting code mentally on every conference
enrichment run. Extracting it makes the output format reproducible and
eliminates the risk of inconsistent column widths or color values.
"""

from __future__ import annotations

from datetime import date
from typing import Any

# ---------------------------------------------------------------------------
# Constants — source: conference-enrichment.md lines 153–158
# ---------------------------------------------------------------------------

#: Header background color — dark navy
HEADER_COLOR: str = "1F4E79"

#: Alternating even-row fill — light blue
ALT_ROW_COLOR: str = "EBF3FB"

#: Column definitions: (header label, field_key, width)
COLUMNS: list[tuple[str, str, int]] = [
    ("Company", "Company", 30),
    ("First Name", "First Name", 18),
    ("Last Name", "Last Name", 18),
    ("Job Title", "Job Title", 35),
    ("LinkedIn", "linkedin", 45),
    ("Email", "email", 35),
]

#: Output filename template — source: conference-enrichment.md line 174
FILENAME_TEMPLATE: str = "{year}_{conference_name}_Pharma_Contacts.xlsx"


# ---------------------------------------------------------------------------
# Builder
# ---------------------------------------------------------------------------


def build_conference_xlsx(
    contacts: list[dict[str, Any]],
    conference_name: str,
    output_path: str | None = None,
) -> str:
    """Build a formatted XLSX for conference pharma contacts.

    Source: conference-enrichment.md lines 143–174.

    Args:
        contacts:       Enriched pharma contacts (list of dicts).
        conference_name: Used in the output filename.
        output_path:    Optional override for output file path.

    Returns:
        Path to the written XLSX file.
    """
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except ImportError as e:
        raise ImportError("openpyxl is required for conference XLSX output") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Pharma Contacts"

    header_fill = PatternFill(
        start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid"
    )
    alt_fill = PatternFill(
        start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type="solid"
    )
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=11)

    # Write headers
    headers = [col[0] for col in COLUMNS]
    for col_idx, header_label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Set column widths — source: conference-enrichment.md lines 158
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row and enable auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Write data rows
    for row_idx, contact in enumerate(contacts, start=2):
        is_even_row = (row_idx % 2 == 0)
        fill = alt_fill if is_even_row else None
        for col_idx, (_, field_key, _) in enumerate(COLUMNS, start=1):
            value = contact.get(field_key, "") or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            if fill:
                cell.fill = fill

    # Determine output path
    if output_path is None:
        year = date.today().year
        safe_name = conference_name.replace(" ", "_").replace("/", "_")
        output_path = FILENAME_TEMPLATE.format(year=year, conference_name=safe_name)

    wb.save(output_path)
    return output_path


def summarize_coverage(contacts: list[dict[str, Any]]) -> dict[str, Any]:
    """Return email and LinkedIn coverage statistics for the delivery summary.

    Source: conference-enrichment.md lines 178–182 (Step 5 summary bullets).
    """
    total = len(contacts)
    with_email = sum(1 for c in contacts if c.get("email"))
    with_linkedin = sum(1 for c in contacts if c.get("linkedin"))
    return {
        "total": total,
        "email_count": with_email,
        "email_pct": round(with_email / total * 100) if total else 0,
        "linkedin_count": with_linkedin,
        "linkedin_pct": round(with_linkedin / total * 100) if total else 0,
    }
